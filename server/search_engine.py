"""

This module provides functionality to parse and search through multiple 
file formats including TXT, HTML, PDF, and XLSX. It supports standard 
string matching, Regular Expressions (regex), and basic Boolean logic (AND/OR).
"""

from __future__ import annotations

import os
import re
from typing import Any, Iterable, List

import openpyxl
import pandas as pd
from PyPDF2 import PdfReader
from bs4 import BeautifulSoup

# --- Configuration ---
# Default directory containing the documents to be indexed
STORAGE_DIR = "storage"
# Number of characters to display around the match for preview
CONTEXT_LENGTH = 30 


def get_context(text: str, index: int, length: int) -> str:
    """
    Extracts a snippet of text around a specific index to provide context.
    
    Args:
        text: The full string to extract from.
        index: The starting position of the match.
        length: How many characters to show before and after.
    """
    start = max(0, index - length)
    end = min(len(text), index + length)
    return text[start:end].strip()


def search_txt(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Searches within a plain text file.
    
    Returns a list of dictionaries containing file info, location, and context.
    """
    results: List[dict[str, Any]] = []
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            for line_num, line in enumerate(f, 1):
                match = re.search(pattern, line) if is_regex else (pattern.lower() in line.lower())
                if match:
                    # Determine start index for context
                    idx = match.start() if is_regex else line.lower().find(pattern.lower())
                    results.append({
                        "file": os.path.basename(filepath),
                        "type": "txt",
                        "location": f"Line {line_num}",
                        "context": get_context(line, idx, CONTEXT_LENGTH),
                    })
    except Exception as e:
        print(f"[ERROR] Failed to read text file {filepath}: {e}")
    return results


def search_html(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Parses and searches within an HTML file using BeautifulSoup to handle tags.
    """
    results: List[dict[str, Any]] = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as file:
            soup = BeautifulSoup(file, "html.parser")
            # Extract text only, ignoring HTML tags
            lines = soup.get_text(separator="\n").splitlines()

        for line_idx, line in enumerate(lines, 1):
            if not line.strip():
                continue

            match = re.search(pattern, line) if is_regex else (pattern.lower() in line.lower())
            if match:
                idx = match.start() if is_regex else line.lower().find(pattern.lower())
                results.append({
                    "file": os.path.basename(filepath),
                    "type": "html",
                    "location": f"Line {line_idx} (Rendered)",
                    "context": get_context(line, idx, CONTEXT_LENGTH),
                })
    except Exception as error:
        print(f"[ERROR] Could not read HTML {filepath}: {error}")
    return results


def search_pdf(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Extracts text from PDF pages and searches for the specified pattern.
    """
    results: List[dict[str, Any]] = []
    try:
        reader = PdfReader(filepath)
        for page_idx, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            if not text:
                continue

            for line in text.splitlines():
                match = re.search(pattern, line) if is_regex else (pattern.lower() in line.lower())
                if match:
                    idx = match.start() if is_regex else line.lower().find(pattern.lower())
                    results.append({
                        "file": os.path.basename(filepath),
                        "type": "pdf",
                        "location": f"Page {page_idx}",
                        "context": get_context(line, idx, CONTEXT_LENGTH),
                    })
    except Exception as error:
        print(f"[ERROR] Could not read PDF {filepath}: {error}")
    return results


def search_excel(filepath: str, pattern: str, is_regex: bool) -> List[dict[str, Any]]:
    """
    Searches through all sheets and cells in an Excel (.xlsx) file.
    """
    results: List[dict[str, Any]] = []
    try:
        workbook = openpyxl.load_workbook(filepath, storage_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    cell_text = str(cell.value)
                    match = re.search(pattern, cell_text) if is_regex else (pattern.lower() in cell_text.lower())
                    
                    if match:
                        idx = match.start() if is_regex else cell_text.lower().find(pattern.lower())
                        results.append({
                            "file": os.path.basename(filepath),
                            "type": "xlsx",
                            "location": f"Sheet '{sheet_name}' | Cell {cell.coordinate}",
                            "context": get_context(cell_text, idx, CONTEXT_LENGTH),
                        })
    except Exception as error:
        print(f"[ERROR] Could not read Excel {filepath}: {error}")
    return results


def _iter_supported_files(directory: str) -> Iterable[tuple[str, str]]:
    """
    Helper generator to yield (filepath, extension) for files in the target directory.
    """
    if not os.path.exists(directory):
        return
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if os.path.isfile(filepath):
            _, ext = os.path.splitext(filename)
            yield filepath, ext.lower()


def process_search(
    query: str, allowed_extensions: list[str], use_regex: bool
) -> List[dict[str, Any]]:
    """
    Main entry point for the search engine. 
    Handles Boolean logic parsing and dispatches tasks to specific file handlers.
    """
    all_results: List[dict[str, Any]] = []

    # --- Logic Parsing ---
    # Split query into terms if Boolean operators are present (case sensitive)
    search_terms: list[str] = [query]
    operator: str | None = None

    if not use_regex:
        if " AND " in query:
            search_terms = [t.strip() for t in query.split(" AND ") if t.strip()]
            operator = "AND"
        elif " OR " in query:
            search_terms = [t.strip() for t in query.split(" OR ") if t.strip()]
            operator = "OR"

    if not os.path.exists(STORAGE_DIR):
        print(f"[WARNING] Documents directory not found: {STORAGE_DIR}")
        return []

    # --- File Iteration & Matching ---
    for filepath, ext in _iter_supported_files(STORAGE_DIR):
        # Filter by user-selected extensions
        if allowed_extensions and ext not in allowed_extensions:
            continue

        # Map extension to the correct parsing function
        handler = None
        if ext == ".txt":
            handler = search_txt
        elif ext == ".html":
            handler = search_html
        elif ext == ".pdf":
            handler = search_pdf
        elif ext in [".xlsx", ".xls"]:
            handler = search_excel

        if handler is None:
            continue

        # Temporary storage for hits in the current file
        per_term_hits: list[List[dict[str, Any]]] = []

        for term in search_terms:
            hits = handler(filepath, term, use_regex)
            if hits:
                per_term_hits.append(hits)
            elif operator == "AND":
                # If one term is missing in AND mode, we can skip this file
                per_term_hits = []
                break

        # Apply Boolean logic to decide which results to keep
        if operator == "AND":
            # Only add if all terms found matches in the file
            if len(per_term_hits) == len(search_terms):
                for hits in per_term_hits:
                    all_results.extend(hits)
        else:
            # Default behavior (OR or single term)
            for hits in per_term_hits:
                all_results.extend(hits)

    return all_results